import os
import sys
import time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))
sys.path.append(str(PROJECT_ROOT / "set12"))

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import skimage.io
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy.ndimage import median_filter
from skimage.metrics import peak_signal_noise_ratio, structural_similarity

from functions.Utils import load_pickle, save_pickle, save_results_to_xlsx
from functions.geonlm_medians_functions import run_geonlm_medians_pipeline
from functions.anlm_functions import run_anlm_pipeline
from NLMedians import run_nlmedians


SOURCE_BASE = Path("/workspace/data/output/set50")
OUTPUT_BASE = Path("/workspace/data/output/set50Hibrid")
LEVELS = ["low", "medium", "moderate", "high", "extreme"]
RESOLUTION = "full_512"

HYBRID_CONFIG = {
    "f": 1,
    "t": 3,
    "nn": 7,
    "h_multiplier": 0.001,
    "switch_impulse_only": True,
    "reject_impulse_candidates": True,
    "use_aswmf_spatial_weights": True,
    "aswmf_weight_diag_1": 1.0,
    "aswmf_weight_diag_2": 1.0,
    "aswmf_weight_other": 10.0,
}

NLMEDIANS_CONFIG = {
    "f": 2,
    "t": 3,
    "h_multiplier": 0.005,
}

MEDIAN_CONFIG = {
    "size": 3,
    "mode": "reflect",
}


def selected_levels():
    raw = os.environ.get("SET50_HIBRID_LEVELS")
    if not raw:
        return LEVELS
    wanted = {part.strip() for part in raw.split(",") if part.strip()}
    return [level for level in LEVELS if level in wanted]


def max_images():
    raw = os.environ.get("SET50_HIBRID_MAX_IMAGES")
    if not raw:
        return None
    return int(raw)


def force_run():
    return os.environ.get("SET50_HIBRID_FORCE", "0") == "1"


def force_hybrid_run():
    return force_run() or os.environ.get("SET50_HIBRID_FORCE_HYBRID", "0") == "1"


def score(psnr, ssim):
    return 0.5 * psnr + 0.5 * (ssim * 100)


def metrics(reference, image):
    reference_uint8 = np.clip(reference, 0, 255).astype(np.uint8)
    image_uint8 = to_grayscale_uint8(image)
    psnr = peak_signal_noise_ratio(reference_uint8, image_uint8, data_range=255)
    ssim = structural_similarity(reference_uint8, image_uint8, data_range=255)
    return psnr, ssim, score(psnr, ssim)


def to_grayscale_uint8(image):
    image = np.asarray(image)
    if image.ndim == 4:
        image = image[0]
    if image.ndim == 3 and image.shape[-1] == 4:
        image = image[..., :3]
    if image.ndim == 3 and image.shape[-1] == 3:
        image = (
            0.2125 * image[..., 0]
            + 0.7154 * image[..., 1]
            + 0.0721 * image[..., 2]
        )
    if image.ndim == 3 and image.shape[0] == 1:
        image = image[0]
    return np.clip(image, 0, 255).astype(np.uint8)


def ensure_dirs(root):
    for subdir in ["NLM", "MEDIAN", "NLMedians", "GEONLMHibrid", "ANLM", "results"]:
        (root / subdir).mkdir(parents=True, exist_ok=True)


def load_original_results_by_file(level):
    path = SOURCE_BASE / f"salt_pepper_{level}" / RESOLUTION / "results" / (
        f"results_salt_pepper_gnlm_median_aswmf_{level}.pkl"
    )
    if not path.exists():
        return {}
    return {row["file_name"]: row for row in load_pickle(path.parent, path.name)}


def save_uint8(path, image):
    skimage.io.imsave(str(path), np.clip(image, 0, 255).astype(np.uint8))


def read_or_run_nlmedians(item, out_path):
    reference = item["img_reference_np"]
    if out_path.exists() and not force_run():
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan

    start = time.time()
    result = run_nlmedians(
        reference=reference,
        noisy=item["img_noisy_salt_pepper_np"],
        h=float(item["nlm_h"]) * NLMEDIANS_CONFIG["h_multiplier"],
        f=NLMEDIANS_CONFIG["f"],
        t=NLMEDIANS_CONFIG["t"],
    )
    elapsed = time.time() - start
    save_uint8(out_path, result["filtered"])
    return result["filtered"], result["psnr"], result["ssim"], result["score"], elapsed


def read_or_run_median(item, out_path):
    reference = item["img_reference_np"]
    if out_path.exists() and not force_run():
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan

    start = time.time()
    filtered = median_filter(
        item["img_noisy_salt_pepper_np"],
        size=MEDIAN_CONFIG["size"],
        mode=MEDIAN_CONFIG["mode"],
    )
    elapsed = time.time() - start
    filtered_uint8 = np.clip(filtered, 0, 255).astype(np.uint8)
    save_uint8(out_path, filtered_uint8)
    psnr, ssim, method_score = metrics(reference, filtered_uint8)
    return filtered_uint8, psnr, ssim, method_score, elapsed


def read_or_run_aswmf_from_folder(item, out_path):
    reference = item["img_reference_np"]
    if out_path.exists():
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score
    return None, np.nan, np.nan, np.nan


def force_anlm_run():
    return force_run() or os.environ.get("SET50_HIBRID_FORCE_ANLM", "0") == "1"


def read_or_run_anlm(item, out_path):
    reference = item["img_reference_np"]
    if out_path.exists() and not force_anlm_run():
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan, (
            float(item["nlm_h"]) * HYBRID_CONFIG["h_multiplier"]
        )

    start = time.time()
    filtered, h_used, psnr, ssim, method_score = run_anlm_pipeline(
        img_original=reference,
        h_base=float(item["nlm_h"]),
        img_noisy=item["img_noisy_salt_pepper_np"],
        f=HYBRID_CONFIG["f"],
        t=HYBRID_CONFIG["t"],
        mult=HYBRID_CONFIG["h_multiplier"],
        switch_impulse_only=HYBRID_CONFIG["switch_impulse_only"],
        reject_impulse_candidates=HYBRID_CONFIG["reject_impulse_candidates"],
        use_aswmf_spatial_weights=HYBRID_CONFIG["use_aswmf_spatial_weights"],
        aswmf_weight_diag_1=HYBRID_CONFIG["aswmf_weight_diag_1"],
        aswmf_weight_diag_2=HYBRID_CONFIG["aswmf_weight_diag_2"],
        aswmf_weight_other=HYBRID_CONFIG["aswmf_weight_other"],
    )
    elapsed = time.time() - start
    save_uint8(out_path, filtered)
    return filtered, psnr, ssim, method_score, elapsed, h_used


def read_or_run_hybrid(item, out_path):
    reference = item["img_reference_np"]
    if out_path.exists() and not force_hybrid_run():
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan, (
            float(item["nlm_h"]) * HYBRID_CONFIG["h_multiplier"]
        )

    start = time.time()
    filtered, h_used, psnr, ssim, method_score = run_geonlm_medians_pipeline(
        img_original=reference,
        h_base=float(item["nlm_h"]),
        img_noisy=item["img_noisy_salt_pepper_np"],
        f=HYBRID_CONFIG["f"],
        t=HYBRID_CONFIG["t"],
        mult=HYBRID_CONFIG["h_multiplier"],
        nn=HYBRID_CONFIG["nn"],
        switch_impulse_only=HYBRID_CONFIG["switch_impulse_only"],
        reject_impulse_candidates=HYBRID_CONFIG["reject_impulse_candidates"],
        use_aswmf_spatial_weights=HYBRID_CONFIG["use_aswmf_spatial_weights"],
        aswmf_weight_diag_1=HYBRID_CONFIG["aswmf_weight_diag_1"],
        aswmf_weight_diag_2=HYBRID_CONFIG["aswmf_weight_diag_2"],
        aswmf_weight_other=HYBRID_CONFIG["aswmf_weight_other"],
    )
    elapsed = time.time() - start
    save_uint8(out_path, filtered)
    return filtered, psnr, ssim, method_score, elapsed, h_used


def warmup_anlm():
    dummy = np.zeros((16, 16), dtype=np.float32)
    dummy[8, 8] = 255.0
    run_anlm_pipeline(
        img_original=dummy,
        h_base=100.0,
        img_noisy=dummy,
        f=HYBRID_CONFIG["f"],
        t=HYBRID_CONFIG["t"],
        mult=HYBRID_CONFIG["h_multiplier"],
        switch_impulse_only=HYBRID_CONFIG["switch_impulse_only"],
        reject_impulse_candidates=HYBRID_CONFIG["reject_impulse_candidates"],
        use_aswmf_spatial_weights=HYBRID_CONFIG["use_aswmf_spatial_weights"],
        aswmf_weight_diag_1=HYBRID_CONFIG["aswmf_weight_diag_1"],
        aswmf_weight_diag_2=HYBRID_CONFIG["aswmf_weight_diag_2"],
        aswmf_weight_other=HYBRID_CONFIG["aswmf_weight_other"],
    )


def run_level(level):
    source_results = SOURCE_BASE / f"salt_pepper_{level}" / RESOLUTION / "results"
    output_root = OUTPUT_BASE / f"salt_pepper_{level}" / RESOLUTION
    output_results = output_root / "results"
    ensure_dirs(output_root)

    nlm_pickle = f"array_nlm_salt_pepper_{level}_filtereds.pkl"
    vector = load_pickle(source_results, nlm_pickle)
    limit = max_images()
    if limit is not None:
        vector = vector[:limit]

    save_pickle(vector, output_results, nlm_pickle)
    original_by_file = load_original_results_by_file(level)
    records = []

    for item in vector:
        file_name = item["file_name"]
        reference = item["img_reference_np"]
        noisy = item["img_noisy_salt_pepper_np"]
        nlm = item["img_filtered_nlm"]

        save_uint8(output_root / "NLM" / file_name, nlm)

        psnr_noisy, ssim_noisy, score_noisy = metrics(reference, noisy)
        psnr_nlm, ssim_nlm, score_nlm = metrics(reference, nlm)

        _, psnr_median, ssim_median, score_median, time_median = read_or_run_median(
            item,
            output_root / "MEDIAN" / file_name,
        )
        _, psnr_nlmed, ssim_nlmed, score_nlmed, time_nlmed = read_or_run_nlmedians(
            item,
            output_root / "NLMedians" / file_name,
        )
        _, psnr_hybrid, ssim_hybrid, score_hybrid, time_hybrid, h_hybrid = read_or_run_hybrid(
            item,
            output_root / "GEONLMHibrid" / file_name,
        )
        _, psnr_anlm, ssim_anlm, score_anlm, time_anlm, h_anlm = read_or_run_anlm(
            item,
            output_root / "ANLM" / file_name,
        )
        _, psnr_aswmf_folder, ssim_aswmf_folder, score_aswmf_folder = read_or_run_aswmf_from_folder(
            item,
            output_root / "ASWMF" / file_name,
        )

        original = original_by_file.get(file_name, {})
        record = {
            "level": level,
            "resolution": RESOLUTION,
            "file_name": file_name,
            "source_nlm_pickle": str(source_results / nlm_pickle),
            "estimated_sigma_salt_pepper": float(item["estimated_sigma_salt_pepper"]),
            "salt_pepper_density": float(item["salt_pepper_density"]),
            "salt_prob": float(item["salt_prob"]),
            "pepper_prob": float(item["pepper_prob"]),
            "psnr_noisy": psnr_noisy,
            "ssim_noisy": ssim_noisy,
            "score_noisy": score_noisy,
            "nlm_h": float(item["nlm_h"]),
            "psnr_nlm": psnr_nlm,
            "ssim_nlm": ssim_nlm,
            "score_nlm": score_nlm,
            "median_size": MEDIAN_CONFIG["size"],
            "median_mode": MEDIAN_CONFIG["mode"],
            "psnr_median": psnr_median,
            "ssim_median": ssim_median,
            "score_median": score_median,
            "time_median": time_median,
            "nlmedians_f": NLMEDIANS_CONFIG["f"],
            "nlmedians_t": NLMEDIANS_CONFIG["t"],
            "nlmedians_h_multiplier": NLMEDIANS_CONFIG["h_multiplier"],
            "nlmedians_h": float(item["nlm_h"]) * NLMEDIANS_CONFIG["h_multiplier"],
            "psnr_nlmedians": psnr_nlmed,
            "ssim_nlmedians": ssim_nlmed,
            "score_nlmedians": score_nlmed,
            "time_nlmedians": time_nlmed,
            "hybrid_f": HYBRID_CONFIG["f"],
            "hybrid_t": HYBRID_CONFIG["t"],
            "hybrid_nn": HYBRID_CONFIG["nn"],
            "hybrid_h_multiplier": HYBRID_CONFIG["h_multiplier"],
            "hybrid_h": h_hybrid,
            "hybrid_switch_impulse_only": HYBRID_CONFIG["switch_impulse_only"],
            "hybrid_reject_impulse_candidates": HYBRID_CONFIG["reject_impulse_candidates"],
            "hybrid_use_aswmf_spatial_weights": HYBRID_CONFIG["use_aswmf_spatial_weights"],
            "hybrid_aswmf_weight_diag_1": HYBRID_CONFIG["aswmf_weight_diag_1"],
            "hybrid_aswmf_weight_diag_2": HYBRID_CONFIG["aswmf_weight_diag_2"],
            "hybrid_aswmf_weight_other": HYBRID_CONFIG["aswmf_weight_other"],
            "psnr_geonlm_hibrid": psnr_hybrid,
            "ssim_geonlm_hibrid": ssim_hybrid,
            "score_geonlm_hibrid": score_hybrid,
            "time_geonlm_hibrid": time_hybrid,
            "anlm_f": HYBRID_CONFIG["f"],
            "anlm_t": HYBRID_CONFIG["t"],
            "anlm_h_multiplier": HYBRID_CONFIG["h_multiplier"],
            "anlm_h": h_anlm,
            "anlm_switch_impulse_only": HYBRID_CONFIG["switch_impulse_only"],
            "anlm_reject_impulse_candidates": HYBRID_CONFIG["reject_impulse_candidates"],
            "anlm_use_aswmf_spatial_weights": HYBRID_CONFIG["use_aswmf_spatial_weights"],
            "psnr_anlm": psnr_anlm,
            "ssim_anlm": ssim_anlm,
            "score_anlm": score_anlm,
            "time_anlm": time_anlm,
            "psnr_aswmf": psnr_aswmf_folder if not np.isnan(psnr_aswmf_folder) else (np.nan if "psnr_aswmf" not in original else float(original["psnr_aswmf"])),
            "ssim_aswmf": ssim_aswmf_folder if not np.isnan(ssim_aswmf_folder) else (np.nan if "ssim_aswmf" not in original else float(original["ssim_aswmf"])),
            "score_aswmf": score_aswmf_folder if not np.isnan(score_aswmf_folder) else (np.nan if "score_aswmf" not in original else float(original["score_aswmf"])),
            "delta_score_hibrid_vs_nlm": score_hybrid - score_nlm,
            "delta_score_anlm_vs_nlm": score_anlm - score_nlm,
            "delta_score_hibrid_vs_anlm": score_hybrid - score_anlm,
            "delta_score_hibrid_vs_median": score_hybrid - score_median,
            "delta_score_hibrid_vs_nlmedians": score_hybrid - score_nlmed,
            "psnr_gnlm_original": np.nan if "psnr_gnlm" not in original else float(original["psnr_gnlm"]),
            "ssim_gnlm_original": np.nan if "ssim_gnlm" not in original else float(original["ssim_gnlm"]),
            "score_gnlm_original": np.nan if "score_gnlm" not in original else float(original["score_gnlm"]),
            "psnr_median_original": np.nan if "psnr_median" not in original else float(original["psnr_median"]),
            "ssim_median_original": np.nan if "ssim_median" not in original else float(original["ssim_median"]),
            "score_median_original": np.nan if "score_median" not in original else float(original["score_median"]),
            "psnr_aswmf_original": np.nan if "psnr_aswmf" not in original else float(original["psnr_aswmf"]),
            "ssim_aswmf_original": np.nan if "ssim_aswmf" not in original else float(original["ssim_aswmf"]),
            "score_aswmf_original": np.nan if "score_aswmf" not in original else float(original["score_aswmf"]),
        }
        records.append(record)
        print(
            f"{level} {file_name}: "
            f"Hibrid={score_hybrid:.4f} "
            f"ANLM={score_anlm:.4f} "
            f"Median={score_median:.4f} "
            f"NLMedians={score_nlmed:.4f} "
            f"dNlMed={record['delta_score_hibrid_vs_nlmedians']:+.4f} "
            f"time={time_hybrid if not np.isnan(time_hybrid) else 0:.2f}s",
            flush=True,
        )

    save_pickle(records, output_results, f"results_set50_hibrid_{level}.pkl")
    save_results_to_xlsx(records, output_results, f"results_set50_hibrid_{level}.xlsx")
    return records


def make_summary(all_records, summary_dir):
    df = pd.DataFrame(all_records)
    summary = (
        df.groupby("level")
        .agg(
            mean_ssim_nlm=("ssim_nlm", "mean"),
            mean_ssim_geonlm_hibrid=("ssim_geonlm_hibrid", "mean"),
            mean_ssim_median=("ssim_median", "mean"),
            mean_ssim_anlm=("ssim_anlm", "mean"),
            mean_ssim_aswmf=("ssim_aswmf", "mean"),
            mean_ssim_nlmedians=("ssim_nlmedians", "mean"),
            mean_psnr_nlm=("psnr_nlm", "mean"),
            mean_psnr_geonlm_hibrid=("psnr_geonlm_hibrid", "mean"),
            mean_psnr_median=("psnr_median", "mean"),
            mean_psnr_anlm=("psnr_anlm", "mean"),
            mean_psnr_aswmf=("psnr_aswmf", "mean"),
            mean_psnr_nlmedians=("psnr_nlmedians", "mean"),
            mean_score_nlm=("score_nlm", "mean"),
            mean_score_geonlm_hibrid=("score_geonlm_hibrid", "mean"),
            mean_score_median=("score_median", "mean"),
            mean_score_anlm=("score_anlm", "mean"),
            mean_score_aswmf=("score_aswmf", "mean"),
            mean_score_nlmedians=("score_nlmedians", "mean"),
            mean_time_geonlm_hibrid=("time_geonlm_hibrid", "mean"),
            mean_time_anlm=("time_anlm", "mean"),
            mean_time_median=("time_median", "mean"),
            mean_time_nlmedians=("time_nlmedians", "mean"),
            wins_hibrid_vs_nlm=("delta_score_hibrid_vs_nlm", lambda s: int((s > 0).sum())),
            wins_anlm_vs_nlm=("delta_score_anlm_vs_nlm", lambda s: int((s > 0).sum())),
            wins_hibrid_vs_anlm=("delta_score_hibrid_vs_anlm", lambda s: int((s > 0).sum())),
            wins_hibrid_vs_median=("delta_score_hibrid_vs_median", lambda s: int((s > 0).sum())),
            wins_hibrid_vs_nlmedians=("delta_score_hibrid_vs_nlmedians", lambda s: int((s > 0).sum())),
        )
        .reset_index()
    )
    summary.to_excel(summary_dir / "results_set50_hibrid_summary.xlsx", index=False)
    return summary


def better(value, reference):
    if pd.isna(value) or pd.isna(reference):
        return ""
    return "melhor" if value > reference else "pior"


def create_method_format_xlsx(all_records, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    header = [
        "",
        "s1 - NLM", "s2 - GNLM", "s3- GHNLM", "s4 - IANLM",
        "s5 - Median", "s6 - ASWMF", "s7 - NLMedian",
        "GHNLM x NLM", "GHNLM x Median", "GHNLM x ASWMF", "GHNLM x NlMedian",
        "p1 - NLM", "p2 - GNLM", "p3- GHNLM", "p4 - IANLM",
        "p5 - Median", "p6 - ASWMF", "p7 - NLMedian",
        "GHNLM x NLM", "GHNLM x Median", "GHNLM x ASWMF", "GHNLM x NlMedian",
        "time_ghnlm", "time_ianlm",
        "nlm_score", "score_gnlm", "score_ghnlm", "score_ianlm",
        "score_median", "score_aswmf", "score_nlmedian",
        "GHNLM x NLM", "GHNLM x Median", "GHNLM x ASWMF", "GHNLM x NlMedian",
        "nlm_h", "h_gnlm", "h_ghnlm", "h_ianlm", "estimated_sigma", "image",
    ]
    for level in LEVELS:
        rows = [row for row in all_records if row["level"] == level]
        if not rows:
            continue
        ws = wb.create_sheet(f"Hibrid Set50 {level.title()}")
        ws.append(header)
        for record in rows:
            ws.append([
                None,
                record["ssim_nlm"],
                record["ssim_gnlm_original"],
                record["ssim_geonlm_hibrid"],
                record["ssim_anlm"],
                record["ssim_median"],
                record["ssim_aswmf"],
                record["ssim_nlmedians"],
                better(record["ssim_geonlm_hibrid"], record["ssim_nlm"]),
                better(record["ssim_geonlm_hibrid"], record["ssim_median"]),
                better(record["ssim_geonlm_hibrid"], record["ssim_aswmf"]),
                better(record["ssim_geonlm_hibrid"], record["ssim_nlmedians"]),
                record["psnr_nlm"],
                record["psnr_gnlm_original"],
                record["psnr_geonlm_hibrid"],
                record["psnr_anlm"],
                record["psnr_median"],
                record["psnr_aswmf"],
                record["psnr_nlmedians"],
                better(record["psnr_geonlm_hibrid"], record["psnr_nlm"]),
                better(record["psnr_geonlm_hibrid"], record["psnr_median"]),
                better(record["psnr_geonlm_hibrid"], record["psnr_aswmf"]),
                better(record["psnr_geonlm_hibrid"], record["psnr_nlmedians"]),
                record["time_geonlm_hibrid"],
                record["time_anlm"],
                record["score_nlm"],
                record["score_gnlm_original"],
                record["score_geonlm_hibrid"],
                record["score_anlm"],
                record["score_median"],
                record["score_aswmf"],
                record["score_nlmedians"],
                better(record["score_geonlm_hibrid"], record["score_nlm"]),
                better(record["score_geonlm_hibrid"], record["score_median"]),
                better(record["score_geonlm_hibrid"], record["score_aswmf"]),
                better(record["score_geonlm_hibrid"], record["score_nlmedians"]),
                record["nlm_h"],
                record.get("h_gnlm_original", np.nan),
                record["hybrid_h"],
                record["anlm_h"],
                record["estimated_sigma_salt_pepper"],
                record["file_name"],
            ])
        style_sheet(ws)
    wb.save(out_path)
    return out_path


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 13
    ws.freeze_panes = "A2"


def add_table_page(pdf, title, df, columns):
    fig, ax = plt.subplots(figsize=(11.69, 8.27))
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=16)
    table_df = df[columns].copy()
    for col in table_df.columns:
        if col != "level":
            table_df[col] = table_df[col].map(lambda x: "" if pd.isna(x) else f"{x:.4f}")
    table = ax.table(
        cellText=table_df.values,
        colLabels=table_df.columns,
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1, 1.35)
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def create_pdfs(summary, out_old, out_new):
    old_cols = [
        "level",
        "mean_score_nlm",
        "mean_score_geonlm_hibrid",
        "mean_score_anlm",
        "mean_score_median",
        "mean_score_aswmf",
        "mean_score_nlmedians",
    ]
    with PdfPages(out_old) as pdf:
        add_table_page(pdf, "Set50 Hibrid - Summary Score", summary, old_cols)

    new_cols_score = [
        "level",
        "mean_score_nlm",
        "mean_score_geonlm_hibrid",
        "mean_score_anlm",
        "mean_score_median",
        "mean_score_aswmf",
        "mean_score_nlmedians",
    ]
    new_cols_psnr = [
        "level",
        "mean_psnr_nlm",
        "mean_psnr_geonlm_hibrid",
        "mean_psnr_anlm",
        "mean_psnr_median",
        "mean_psnr_aswmf",
        "mean_psnr_nlmedians",
    ]
    new_cols_ssim = [
        "level",
        "mean_ssim_nlm",
        "mean_ssim_geonlm_hibrid",
        "mean_ssim_anlm",
        "mean_ssim_median",
        "mean_ssim_aswmf",
        "mean_ssim_nlmedians",
    ]
    with PdfPages(out_new) as pdf:
        add_table_page(pdf, "Set50 Hibrid - Score Format", summary, new_cols_score)
        add_table_page(pdf, "Set50 Hibrid - PSNR Format", summary, new_cols_psnr)
        add_table_page(pdf, "Set50 Hibrid - SSIM Format", summary, new_cols_ssim)


if __name__ == "__main__":
    warmup_anlm()
    all_records = []
    for level in selected_levels():
        all_records.extend(run_level(level))

    summary_dir = OUTPUT_BASE / "summary"
    summary_dir.mkdir(parents=True, exist_ok=True)
    save_pickle(all_records, summary_dir, "results_set50_hibrid_all.pkl")
    save_results_to_xlsx(all_records, summary_dir, "results_set50_hibrid_all.xlsx")

    summary = make_summary(all_records, summary_dir)
    method_xlsx = create_method_format_xlsx(
        all_records,
        summary_dir / "Results Salt and Pepper Hibrid Method - Set50.xlsx",
    )
    create_pdfs(
        summary,
        summary_dir / "results_set50_hibrid_summary_old_format.pdf",
        summary_dir / "results_set50_hibrid_method_format.pdf",
    )
    print(summary.to_string(index=False))
    print(f"Method-format XLSX saved to: {method_xlsx}")
