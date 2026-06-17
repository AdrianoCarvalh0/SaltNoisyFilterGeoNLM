from pathlib import Path
import pickle

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_DIR = Path("/workspace/data/output/hibrid_statistical_tables")
DATASETS = [
    ("set12", "Set12"),
    ("set50", "Set50"),
]
LEVEL_ORDER = ["low", "medium", "moderate", "high", "extreme"]

METHODS = [
    ("NLM", "nlm"),
    ("GNLM", "gnlm"),
    ("GHNLM", "geonlm_hibrid"),
    ("NLM+ASWMF", "aswmf_nlm_h001"),
    ("Median", "median"),
    ("ASWMF", "aswmf"),
    ("NLMedian", "nlmedians"),
]

METRICS = [
    ("SSIM", "ssim"),
    ("PSNR", "psnr"),
    ("Score", "score"),
]


def load_dataset(dataset_name, dataset_label):
    path = Path(f"/workspace/data/output/{dataset_name}Hibrid/summary/results_{dataset_name}_hibrid_all.pkl")
    with path.open("rb") as file:
        rows = pickle.load(file)
    df = pd.DataFrame(rows)
    df["dataset"] = dataset_label
    df["level"] = pd.Categorical(df["level"], categories=LEVEL_ORDER, ordered=True)
    return df.sort_values(["level", "file_name"]).reset_index(drop=True)


def method_column(metric, method_key):
    if method_key == "aswmf":
        return f"{metric}_aswmf_original"
    return f"{metric}_{method_key}"


def individual_metric_table(df, metric):
    data = {
        "dataset": df["dataset"],
        "level": df["level"].astype(str),
        "image": df["file_name"],
    }
    method_columns = []
    for label, key in METHODS:
        column = method_column(metric, key)
        data[label] = df[column]
        method_columns.append(label)

    table = pd.DataFrame(data)
    values = table[method_columns].to_numpy(dtype=float)
    winners = np.nanargmax(values, axis=1)
    table["winner"] = [method_columns[index] for index in winners]
    return table


def ghnlm_win_rate(df, metric):
    table = individual_metric_table(df, metric)
    method_labels = [label for label, _ in METHODS]
    values = table[method_labels].to_numpy(dtype=float)
    ghnlm = table["GHNLM"].to_numpy(dtype=float)
    wins = int(np.sum(ghnlm >= np.nanmax(values, axis=1)))
    total = int(len(table))
    return wins, total, 100.0 * wins / total


def ghnlm_statistical_summary(df):
    rows = []
    for metric_label, metric in METRICS:
        wins, total, pct = ghnlm_win_rate(df, metric)
        rows.append({
            "Metrica": metric_label,
            "Media GHNLM": df[method_column(metric, "geonlm_hibrid")].mean(),
            "Win-rate GHNLM": f"{wins}/{total} = {pct:.1f}%",
            "Wins": wins,
            "Total": total,
            "Win-rate (%)": pct,
        })
    return pd.DataFrame(rows)


def method_means(df):
    rows = []
    for label, key in METHODS:
        rows.append({
            "Metodo": label,
            "SSIM medio": df[method_column("ssim", key)].mean(),
            "PSNR medio": df[method_column("psnr", key)].mean(),
            "Score medio": df[method_column("score", key)].mean(),
        })
    return pd.DataFrame(rows)


def dataset_method_means(all_df):
    chunks = []
    for dataset_label, group in all_df.groupby("dataset", sort=False):
        table = method_means(group)
        table.insert(0, "Dataset", dataset_label)
        chunks.append(table)
    return pd.concat(chunks, ignore_index=True)


def format_workbook(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                if isinstance(cell.value, float):
                    if "SSIM" in str(ws.cell(1, cell.column).value):
                        cell.number_format = "0.0000"
                    elif "%" in str(ws.cell(1, cell.column).value):
                        cell.number_format = "0.0"
                    else:
                        cell.number_format = "0.00"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15
        ws.freeze_panes = "A2"
    wb.save(path)


def save_xlsx(all_df, per_dataset):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / "hibrid_statistical_tables.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overall_win = ghnlm_statistical_summary(all_df)
        overall_means = method_means(all_df)
        dataset_means = dataset_method_means(all_df)

        overall_win.to_excel(writer, sheet_name="Resumo GHNLM geral", index=False)
        overall_means.to_excel(writer, sheet_name="Medias gerais metodos", index=False)
        dataset_means.to_excel(writer, sheet_name="Medias por dataset", index=False)

        for dataset_label, df in per_dataset.items():
            ghnlm_statistical_summary(df).to_excel(
                writer,
                sheet_name=f"{dataset_label} resumo GHNLM",
                index=False,
            )
            method_means(df).to_excel(
                writer,
                sheet_name=f"{dataset_label} medias",
                index=False,
            )
            for metric_label, metric in METRICS:
                individual_metric_table(df, metric).to_excel(
                    writer,
                    sheet_name=f"{dataset_label} {metric_label}",
                    index=False,
                )

    format_workbook(path)
    return path


def add_table_page(pdf, title, df, font_size=8, scale_y=1.25):
    fig_width = max(11.69, min(22, 1.35 * len(df.columns)))
    fig, ax = plt.subplots(figsize=(fig_width, 8.27))
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=16)

    table_df = df.copy()
    for column in table_df.columns:
        if pd.api.types.is_numeric_dtype(table_df[column]):
            if "SSIM" in column:
                table_df[column] = table_df[column].map(lambda value: "" if pd.isna(value) else f"{value:.4f}")
            elif "%" in column:
                table_df[column] = table_df[column].map(lambda value: "" if pd.isna(value) else f"{value:.1f}")
            else:
                table_df[column] = table_df[column].map(lambda value: "" if pd.isna(value) else f"{value:.2f}")

    table = ax.table(
        cellText=table_df.values,
        colLabels=table_df.columns,
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(font_size)
    table.scale(1, scale_y)
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def add_paginated_table(pdf, title, df, rows_per_page=22):
    for start in range(0, len(df), rows_per_page):
        page = df.iloc[start:start + rows_per_page]
        page_number = start // rows_per_page + 1
        total_pages = int(np.ceil(len(df) / rows_per_page))
        add_table_page(
            pdf,
            f"{title} ({page_number}/{total_pages})",
            page,
            font_size=6,
            scale_y=1.05,
        )


def save_pdf(all_df, per_dataset):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / "hibrid_statistical_tables.pdf"
    with PdfPages(path) as pdf:
        add_table_page(
            pdf,
            "Resumo estatistico geral - GHNLM",
            ghnlm_statistical_summary(all_df)[["Metrica", "Media GHNLM", "Win-rate GHNLM"]],
            font_size=10,
            scale_y=1.5,
        )
        add_table_page(
            pdf,
            "Medias gerais por metodo",
            method_means(all_df),
            font_size=9,
            scale_y=1.35,
        )
        add_table_page(
            pdf,
            "Medias por dataset e metodo",
            dataset_method_means(all_df),
            font_size=8,
            scale_y=1.15,
        )
        for dataset_label, df in per_dataset.items():
            add_table_page(
                pdf,
                f"{dataset_label} - Resumo GHNLM",
                ghnlm_statistical_summary(df)[["Metrica", "Media GHNLM", "Win-rate GHNLM"]],
                font_size=10,
                scale_y=1.5,
            )
            add_table_page(
                pdf,
                f"{dataset_label} - Medias por metodo",
                method_means(df),
                font_size=9,
                scale_y=1.35,
            )
            for metric_label, metric in METRICS:
                add_paginated_table(
                    pdf,
                    f"{dataset_label} - Resultados individuais {metric_label}",
                    individual_metric_table(df, metric),
                    rows_per_page=24 if dataset_label == "Set12" else 22,
                )
    return path


def main():
    per_dataset = {
        dataset_label: load_dataset(dataset_name, dataset_label)
        for dataset_name, dataset_label in DATASETS
    }
    all_df = pd.concat(per_dataset.values(), ignore_index=True)

    xlsx_path = save_xlsx(all_df, per_dataset)
    pdf_path = save_pdf(all_df, per_dataset)

    print(xlsx_path)
    print(pdf_path)
    print(ghnlm_statistical_summary(all_df)[["Metrica", "Media GHNLM", "Win-rate GHNLM"]].to_string(index=False))
    print(method_means(all_df).to_string(index=False))


if __name__ == "__main__":
    main()
