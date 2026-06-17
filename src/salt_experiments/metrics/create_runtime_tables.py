from pathlib import Path
import pickle
import time

import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
from scipy.ndimage import median_filter
from openpyxl.styles import Alignment, Font, PatternFill

import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

from functions.NLMedians import run_nlmedians
from functions.salt_filters import aswmf_filter


OUTPUT_DIR = Path("/workspace/data/output/hibrid_runtime_tables")
LEVELS = ["low", "moderate", "medium", "high", "extreme"]
DENSITY = {
    "low": 0.01,
    "moderate": 0.03,
    "medium": 0.05,
    "high": 0.10,
    "extreme": 0.20,
}
COMMENTS = {
    "low": "Low impulse density",
    "moderate": "Moderate impulse density",
    "medium": "Medium impulse density",
    "high": "High impulse density",
    "extreme": "Extreme impulse density",
}
DATASETS = [("set12", "Set12"), ("set50", "Set50")]
NL_MEDIAN_SAMPLE_PER_LEVEL = 3


def load_hibrid_all(dataset_name):
    path = Path(f"/workspace/data/output/{dataset_name}Hibrid/summary/results_{dataset_name}_hibrid_all.pkl")
    with path.open("rb") as file:
        return pd.DataFrame(pickle.load(file))


def load_nlm_vector(dataset_name, level):
    path = (
        Path(f"/workspace/data/output/{dataset_name}Hibrid")
        / f"salt_pepper_{level}"
        / "full_512"
        / "results"
        / f"array_nlm_salt_pepper_{level}_filtereds.pkl"
    )
    with path.open("rb") as file:
        return pickle.load(file)


def load_nlm_cuda_runtime(dataset_label, level):
    path = OUTPUT_DIR / "nlm_cuda_runtime_details.pkl"
    if not path.exists():
        return np.nan, 0
    with path.open("rb") as file:
        df = pd.DataFrame(pickle.load(file))
    subset = df[(df["dataset"] == dataset_label) & (df["level"] == level)]
    if subset.empty:
        return np.nan, 0
    return float(subset["runtime_s"].mean()), int(len(subset))


def warm_up_numba(sample):
    noisy = sample["img_noisy_salt_pepper_np"].astype(np.float32)
    crop = noisy[:32, :32]
    aswmf_filter(crop, radius=3)
    run_nlmedians(crop, crop, h=1.0, f=2, t=3)


def measure_median_aswmf(vector):
    median_times = []
    aswmf_times = []
    for item in vector:
        noisy = item["img_noisy_salt_pepper_np"].astype(np.float32)

        start = time.perf_counter()
        median_filter(noisy, size=3, mode="reflect")
        median_times.append(time.perf_counter() - start)

        start = time.perf_counter()
        aswmf_filter(
            noisy,
            radius=3,
            weight_diag_1=1.0,
            weight_diag_2=1.0,
            weight_other=10.0,
        )
        aswmf_times.append(time.perf_counter() - start)

    return np.array(median_times), np.array(aswmf_times)


def measure_nlmedian_sample(vector, sample_size=NL_MEDIAN_SAMPLE_PER_LEVEL):
    times = []
    for item in vector[:sample_size]:
        start = time.perf_counter()
        run_nlmedians(
            reference=item["img_reference_np"],
            noisy=item["img_noisy_salt_pepper_np"],
            h=float(item["nlm_h"]) * 0.005,
            f=2,
            t=3,
        )
        times.append(time.perf_counter() - start)
    return np.array(times)


def safe_mean(series):
    series = pd.to_numeric(series, errors="coerce")
    if series.notna().sum() == 0:
        return np.nan
    return float(series.mean())


def build_dataset_runtime(dataset_name, dataset_label):
    hibrid = load_hibrid_all(dataset_name)
    records = []
    detail_records = []

    first_vector = load_nlm_vector(dataset_name, "low")
    warm_up_numba(first_vector[0])

    for level in LEVELS:
        level_rows = hibrid[hibrid["level"] == level]
        vector = load_nlm_vector(dataset_name, level)

        median_times, aswmf_times = measure_median_aswmf(vector)
        nlmedian_times = measure_nlmedian_sample(vector)
        nlm_cuda_mean, nlm_cuda_n = load_nlm_cuda_runtime(dataset_label, level)

        for index, elapsed in enumerate(median_times):
            detail_records.append({
                "dataset": dataset_label,
                "level": level,
                "method": "Median",
                "sample_index": index,
                "runtime_s": elapsed,
            })
        for index, elapsed in enumerate(aswmf_times):
            detail_records.append({
                "dataset": dataset_label,
                "level": level,
                "method": "ASWMF",
                "sample_index": index,
                "runtime_s": elapsed,
            })
            detail_records.append({
                "dataset": dataset_label,
                "level": level,
                "method": "ANLM",
                "sample_index": index,
                "runtime_s": elapsed,
            })
        for index, elapsed in enumerate(nlmedian_times):
            detail_records.append({
                "dataset": dataset_label,
                "level": level,
                "method": "NLMedian",
                "sample_index": index,
                "runtime_s": elapsed,
            })

        records.append({
            "Dataset": dataset_label,
            "Noise density": f"p={DENSITY[level]:.2f}",
            "Level": level.title(),
            "NLM (s)": nlm_cuda_mean,
            "GNLM (s)": safe_mean(level_rows["time_geonlm"]),
            "GHNLM (s)": safe_mean(level_rows["time_geonlm_hibrid"]),
            "ANLM (s)": float(aswmf_times.mean()),
            "Median (s)": float(median_times.mean()),
            "ASWMF (s)": float(aswmf_times.mean()),
            "NLMedian (s)": float(nlmedian_times.mean()),
            "NLM n": nlm_cuda_n,
            "GNLM n": int(level_rows["time_geonlm"].notna().sum()),
            "GHNLM n": int(level_rows["time_geonlm_hibrid"].notna().sum()),
            "Median n": int(len(median_times)),
            "ASWMF n": int(len(aswmf_times)),
            "ANLM n": int(len(aswmf_times)),
            "NLMedian n": int(len(nlmedian_times)),
            "Comment": COMMENTS[level],
        })

    return pd.DataFrame(records), pd.DataFrame(detail_records)


def format_workbook(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15
        ws.freeze_panes = "A2"
    wb.save(path)


def save_xlsx(runtime_tables, detail_tables):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / "runtime_tables.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        all_runtime = pd.concat(runtime_tables, ignore_index=True)
        all_details = pd.concat(detail_tables, ignore_index=True)
        all_runtime.to_excel(writer, sheet_name="Runtime summary", index=False)
        for dataset_label, table in zip([label for _, label in DATASETS], runtime_tables):
            table.to_excel(writer, sheet_name=f"{dataset_label} runtime", index=False)
        all_details.to_excel(writer, sheet_name="Measured details", index=False)
    format_workbook(path)
    return path


def add_table_page(pdf, title, table):
    fig, ax = plt.subplots(figsize=(13.5, 7.8))
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=16)

    display = table.copy()
    for column in display.columns:
        if column.endswith("(s)"):
            display[column] = display[column].map(lambda value: "--" if pd.isna(value) else f"{value:.2f}")

    table_artist = ax.table(
        cellText=display.values,
        colLabels=display.columns,
        loc="center",
        cellLoc="center",
    )
    table_artist.auto_set_font_size(False)
    table_artist.set_fontsize(7)
    table_artist.scale(1, 1.28)
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def save_pdf(runtime_tables):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / "runtime_tables.pdf"
    with PdfPages(path) as pdf:
        for (_, dataset_label), table in zip(DATASETS, runtime_tables):
            columns = [
                "Noise density",
                "NLM (s)",
                "GNLM (s)",
                "GHNLM (s)",
                "ANLM (s)",
                "Median (s)",
                "ASWMF (s)",
                "NLMedian (s)",
                "Comment",
            ]
            add_table_page(pdf, f"Average runtime on {dataset_label}", table[columns])
    return path


def latex_table(dataset_label, table):
    columns = [
        ("Noise density", None),
        ("GNLM", "GNLM (s)"),
        ("GHNLM", "GHNLM (s)"),
        ("ANLM", "ANLM (s)"),
        ("Median", "Median (s)"),
        ("ASWMF", "ASWMF (s)"),
        ("NLMedian", "NLMedian (s)"),
        ("Comment", None),
    ]
    lines = [
        "\\begin{table}[!t]",
        "\\centering",
        f"\\caption{{Average runtime of the evaluated methods on {dataset_label}.}}",
        f"\\label{{tab:runtime-{dataset_label.lower()}}}",
        "\\scriptsize",
        "\\begin{tabular}{lcccccccc}",
        "\\toprule",
        "Noise density & NLM & GNLM & GHNLM & ANLM & Median & ASWMF & NLMedian & Comment \\\\",
        "\\midrule",
    ]
    for _, row in table.iterrows():
        density = row["Noise density"].replace("p=", "\\(p=") + "\\)"
        values = []
        for column in ["NLM (s)", "GNLM (s)", "GHNLM (s)", "ANLM (s)", "Median (s)", "ASWMF (s)", "NLMedian (s)"]:
            value = row[column]
            values.append("--" if pd.isna(value) else f"{value:.2f}")
        lines.append(
            f"{density} & "
            + " & ".join(values)
            + f" & {row['Comment']} \\\\"
        )
    lines.extend([
        "\\bottomrule",
        "\\end{tabular}",
        "\\end{table}",
    ])
    return "\n".join(lines)


def save_latex(runtime_tables):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    paths = []
    for (_, dataset_label), table in zip(DATASETS, runtime_tables):
        path = OUTPUT_DIR / f"runtime_table_{dataset_label.lower()}.tex"
        path.write_text(latex_table(dataset_label, table), encoding="utf-8")
        paths.append(path)
    return paths


def main():
    runtime_tables = []
    detail_tables = []
    for dataset_name, dataset_label in DATASETS:
        runtime, details = build_dataset_runtime(dataset_name, dataset_label)
        runtime_tables.append(runtime)
        detail_tables.append(details)

    xlsx_path = save_xlsx(runtime_tables, detail_tables)
    pdf_path = save_pdf(runtime_tables)
    tex_paths = save_latex(runtime_tables)

    print(xlsx_path)
    print(pdf_path)
    for path in tex_paths:
        print(path)
    print(pd.concat(runtime_tables, ignore_index=True).to_string(index=False))


if __name__ == "__main__":
    main()
