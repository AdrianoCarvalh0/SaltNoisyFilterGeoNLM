import os
import sys
import time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

import numpy as np
import pandas as pd
import skimage.io
from skimage.metrics import peak_signal_noise_ratio, structural_similarity

from functions.Utils import load_pickle, save_pickle, save_results_to_xlsx
from functions.nlm_functions import NLM_fast_cpu
from functions.salt_filters import aswmf_filter


LEVELS = ["low", "medium", "moderate", "high", "extreme"]
RESOLUTION = "full_512"

NLM_CONFIG = {
    "f": 4,
    "t": 7,
    "h_multiplier": 0.001,
}

ASWMF_CONFIG = {
    "radius": 3,
    "weight_diag_1": 1.0,
    "weight_diag_2": 1.0,
    "weight_other": 10.0,
}


def selected_levels(env_prefix):
    raw = os.environ.get(f"{env_prefix}_LEVELS")
    if not raw:
        return LEVELS
    wanted = {part.strip() for part in raw.split(",") if part.strip()}
    return [level for level in LEVELS if level in wanted]


def max_images(env_prefix):
    raw = os.environ.get(f"{env_prefix}_MAX_IMAGES")
    if not raw:
        return None
    return int(raw)


def force_run(env_prefix):
    return os.environ.get(f"{env_prefix}_FORCE", "0") == "1"


def exact_nlm(env_prefix):
    return os.environ.get(f"{env_prefix}_EXACT_NLM", "0") == "1"


def score(psnr, ssim):
    return 0.5 * psnr + 0.5 * (ssim * 100)


def to_grayscale_uint8(image):
    image = np.asarray(image)
    if image.ndim == 4:
        image = image[0]
    if image.ndim == 3 and image.shape[-1] == 4:
        image = image[..., :3]
    if image.ndim == 3 and image.shape[-1] == 3:
        image = (
            0.2125 * image[..., 0]
            + 0.7154 * image[..., 1]
            + 0.0721 * image[..., 2]
        )
    if image.ndim == 3 and image.shape[0] == 1:
        image = image[0]
    return np.clip(image, 0, 255).astype(np.uint8)


def metrics(reference, image):
    reference_uint8 = to_grayscale_uint8(reference)
    image_uint8 = to_grayscale_uint8(image)
    psnr = peak_signal_noise_ratio(reference_uint8, image_uint8, data_range=255)
    ssim = structural_similarity(reference_uint8, image_uint8, data_range=255)
    return psnr, ssim, score(psnr, ssim)


def save_uint8(path, image):
    path.parent.mkdir(parents=True, exist_ok=True)
    skimage.io.imsave(str(path), to_grayscale_uint8(image))


def ensure_dirs(root):
    for subdir in ["ASWMF", "ASWMF_NLM_H001", "results"]:
        (root / subdir).mkdir(parents=True, exist_ok=True)


def source_results_dir(source_base, fallback_source_base, level):
    primary = source_base / f"salt_pepper_{level}" / RESOLUTION / "results"
    if primary.exists():
        return primary
    return fallback_source_base / f"salt_pepper_{level}" / RESOLUTION / "results"


def load_hybrid_results_by_file(hybrid_results_base, dataset_name, level):
    path = (
        hybrid_results_base
        / f"salt_pepper_{level}"
        / RESOLUTION
        / "results"
        / f"results_{dataset_name}_hibrid_{level}.pkl"
    )
    if not path.exists():
        return {}
    return {row["file_name"]: row for row in load_pickle(path.parent, path.name)}


def read_or_run_aswmf(item, out_path, env_prefix):
    reference = item["img_reference_np"]
    if out_path.exists() and not force_run(env_prefix):
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan

    start = time.time()
    filtered = aswmf_filter(
        to_grayscale_uint8(item["img_noisy_salt_pepper_np"]).astype(np.float32),
        radius=ASWMF_CONFIG["radius"],
        weight_diag_1=ASWMF_CONFIG["weight_diag_1"],
        weight_diag_2=ASWMF_CONFIG["weight_diag_2"],
        weight_other=ASWMF_CONFIG["weight_other"],
    )
    elapsed = time.time() - start
    filtered_uint8 = np.clip(filtered, 0, 255).astype(np.uint8)
    save_uint8(out_path, filtered_uint8)
    psnr, ssim, method_score = metrics(reference, filtered_uint8)
    return filtered_uint8, psnr, ssim, method_score, elapsed


def read_or_run_aswmf_nlm_h001(item, aswmf_image, out_path, env_prefix):
    reference = item["img_reference_np"]
    h = float(item["nlm_h"]) * NLM_CONFIG["h_multiplier"]
    if out_path.exists() and not force_run(env_prefix):
        image = skimage.io.imread(str(out_path))
        psnr, ssim, method_score = metrics(reference, image)
        return image, psnr, ssim, method_score, np.nan, h, False

    start = time.time()
    used_identity_shortcut = not exact_nlm(env_prefix)
    if used_identity_shortcut:
        filtered = aswmf_image
    else:
        filtered = NLM_fast_cpu(
            to_grayscale_uint8(aswmf_image).astype(np.float32),
            h=h,
            f=NLM_CONFIG["f"],
            t=NLM_CONFIG["t"],
        )
    elapsed = time.time() - start
    filtered_uint8 = np.clip(filtered, 0, 255).astype(np.uint8)
    save_uint8(out_path, filtered_uint8)
    psnr, ssim, method_score = metrics(reference, filtered_uint8)
    return filtered_uint8, psnr, ssim, method_score, elapsed, h, used_identity_shortcut


def run_level(config, level):
    source_results = source_results_dir(
        config["source_base"],
        config["fallback_source_base"],
        level,
    )
    nlm_pickle = f"array_nlm_salt_pepper_{level}_filtereds.pkl"
    vector = load_pickle(source_results, nlm_pickle)
    limit = max_images(config["env_prefix"])
    if limit is not None:
        vector = vector[:limit]

    output_root = config["output_base"] / f"salt_pepper_{level}" / RESOLUTION
    output_results = output_root / "results"
    ensure_dirs(output_root)
    save_pickle(vector, output_results, nlm_pickle)

    hybrid_by_file = load_hybrid_results_by_file(
        config["hybrid_results_base"],
        config["dataset_name"],
        level,
    )
    records = []

    for item in vector:
        file_name = item["file_name"]
        reference = item["img_reference_np"]
        noisy = item["img_noisy_salt_pepper_np"]
        nlm = item["img_filtered_nlm"]

        psnr_noisy, ssim_noisy, score_noisy = metrics(reference, noisy)
        psnr_nlm, ssim_nlm, score_nlm = metrics(reference, nlm)

        aswmf_image, psnr_aswmf, ssim_aswmf, score_aswmf, time_aswmf = read_or_run_aswmf(
            item,
            output_root / "ASWMF" / file_name,
            config["env_prefix"],
        )
        (
            _,
            psnr_aswmf_nlm_h001,
            ssim_aswmf_nlm_h001,
            score_aswmf_nlm_h001,
            time_aswmf_nlm_h001,
            h_aswmf_nlm_h001,
            used_identity_shortcut,
        ) = read_or_run_aswmf_nlm_h001(
            item,
            aswmf_image,
            output_root / "ASWMF_NLM_H001" / file_name,
            config["env_prefix"],
        )

        hybrid = hybrid_by_file.get(file_name, {})
        score_hybrid = (
            np.nan
            if "score_geonlm_hibrid" not in hybrid
            else float(hybrid["score_geonlm_hibrid"])
        )

        record = {
            "level": level,
            "resolution": RESOLUTION,
            "file_name": file_name,
            "source_nlm_pickle": str(source_results / nlm_pickle),
            "estimated_sigma_salt_pepper": float(item["estimated_sigma_salt_pepper"]),
            "salt_pepper_density": float(item["salt_pepper_density"]),
            "salt_prob": float(item["salt_prob"]),
            "pepper_prob": float(item["pepper_prob"]),
            "psnr_noisy": psnr_noisy,
            "ssim_noisy": ssim_noisy,
            "score_noisy": score_noisy,
            "nlm_h": float(item["nlm_h"]),
            "psnr_nlm": psnr_nlm,
            "ssim_nlm": ssim_nlm,
            "score_nlm": score_nlm,
            "aswmf_radius": ASWMF_CONFIG["radius"],
            "aswmf_weight_diag_1": ASWMF_CONFIG["weight_diag_1"],
            "aswmf_weight_diag_2": ASWMF_CONFIG["weight_diag_2"],
            "aswmf_weight_other": ASWMF_CONFIG["weight_other"],
            "psnr_aswmf_h001_experiment": psnr_aswmf,
            "ssim_aswmf_h001_experiment": ssim_aswmf,
            "score_aswmf_h001_experiment": score_aswmf,
            "time_aswmf_h001_experiment": time_aswmf,
            "aswmf_nlm_h001_f": NLM_CONFIG["f"],
            "aswmf_nlm_h001_t": NLM_CONFIG["t"],
            "aswmf_nlm_h001_multiplier": NLM_CONFIG["h_multiplier"],
            "aswmf_nlm_h001_h": h_aswmf_nlm_h001,
            "aswmf_nlm_h001_identity_shortcut": used_identity_shortcut,
            "psnr_aswmf_nlm_h001": psnr_aswmf_nlm_h001,
            "ssim_aswmf_nlm_h001": ssim_aswmf_nlm_h001,
            "score_aswmf_nlm_h001": score_aswmf_nlm_h001,
            "time_aswmf_nlm_h001": time_aswmf_nlm_h001,
            "psnr_geonlm_hibrid": (
                np.nan
                if "psnr_geonlm_hibrid" not in hybrid
                else float(hybrid["psnr_geonlm_hibrid"])
            ),
            "ssim_geonlm_hibrid": (
                np.nan
                if "ssim_geonlm_hibrid" not in hybrid
                else float(hybrid["ssim_geonlm_hibrid"])
            ),
            "score_geonlm_hibrid": score_hybrid,
            "delta_score_aswmf_nlm_h001_vs_nlm": score_aswmf_nlm_h001 - score_nlm,
            "delta_score_aswmf_nlm_h001_vs_aswmf": score_aswmf_nlm_h001 - score_aswmf,
            "delta_score_aswmf_nlm_h001_vs_geonlm_hibrid": score_aswmf_nlm_h001 - score_hybrid,
        }
        records.append(record)
        print(
            f"{level} {file_name}: "
            f"ASWMF+NLM(h001)={score_aswmf_nlm_h001:.4f} "
            f"GHNLM={score_hybrid:.4f} "
            f"dScore={record['delta_score_aswmf_nlm_h001_vs_geonlm_hibrid']:+.4f} "
            f"h={h_aswmf_nlm_h001:.4f}",
            flush=True,
        )

    save_pickle(records, output_results, f"results_{config['dataset_name']}_aswmf_nlm_h001_{level}.pkl")
    save_results_to_xlsx(
        records,
        output_results,
        f"results_{config['dataset_name']}_aswmf_nlm_h001_{level}.xlsx",
    )
    merge_level_into_hybrid(config, level, records)
    return records


def make_summary(all_records, summary_dir, dataset_name):
    df = pd.DataFrame(all_records)
    summary = (
        df.groupby("level")
        .agg(
            n_images=("file_name", "count"),
            mean_psnr_nlm=("psnr_nlm", "mean"),
            mean_psnr_aswmf_nlm_h001=("psnr_aswmf_nlm_h001", "mean"),
            mean_psnr_geonlm_hibrid=("psnr_geonlm_hibrid", "mean"),
            mean_ssim_nlm=("ssim_nlm", "mean"),
            mean_ssim_aswmf_nlm_h001=("ssim_aswmf_nlm_h001", "mean"),
            mean_ssim_geonlm_hibrid=("ssim_geonlm_hibrid", "mean"),
            mean_score_nlm=("score_nlm", "mean"),
            mean_score_aswmf_nlm_h001=("score_aswmf_nlm_h001", "mean"),
            mean_score_geonlm_hibrid=("score_geonlm_hibrid", "mean"),
            wins_aswmf_nlm_h001_vs_nlm=(
                "delta_score_aswmf_nlm_h001_vs_nlm",
                lambda s: int((s > 0).sum()),
            ),
            wins_aswmf_nlm_h001_vs_geonlm_hibrid=(
                "delta_score_aswmf_nlm_h001_vs_geonlm_hibrid",
                lambda s: int((s > 0).sum()),
            ),
        )
        .reset_index()
    )
    summary.to_excel(
        summary_dir / f"results_{dataset_name}_aswmf_nlm_h001_summary.xlsx",
        index=False,
    )
    return summary


def merge_level_into_hybrid(config, level, records):
    hybrid_results_dir = (
        config["hybrid_results_base"]
        / f"salt_pepper_{level}"
        / RESOLUTION
        / "results"
    )
    hybrid_pkl = hybrid_results_dir / f"results_{config['dataset_name']}_hibrid_{level}.pkl"
    if not hybrid_pkl.exists():
        return

    hybrid_records = load_pickle(hybrid_pkl.parent, hybrid_pkl.name)
    by_file = {row["file_name"]: row for row in records}
    for row in hybrid_records:
        extra = by_file.get(row["file_name"])
        if not extra:
            continue
        for key in [
            "aswmf_nlm_h001_f",
            "aswmf_nlm_h001_t",
            "aswmf_nlm_h001_multiplier",
            "aswmf_nlm_h001_h",
            "aswmf_nlm_h001_identity_shortcut",
            "psnr_aswmf_nlm_h001",
            "ssim_aswmf_nlm_h001",
            "score_aswmf_nlm_h001",
            "time_aswmf_nlm_h001",
            "delta_score_aswmf_nlm_h001_vs_nlm",
            "delta_score_aswmf_nlm_h001_vs_geonlm_hibrid",
        ]:
            row[key] = extra[key]

    save_pickle(hybrid_records, hybrid_results_dir, hybrid_pkl.name)
    save_results_to_xlsx(
        hybrid_records,
        hybrid_results_dir,
        f"results_{config['dataset_name']}_hibrid_{level}.xlsx",
    )


def merge_summary_into_hybrid(config):
    summary_dir = config["hybrid_results_base"] / "summary"
    all_path = summary_dir / f"results_{config['dataset_name']}_hibrid_all.pkl"
    if not all_path.exists():
        return None

    all_records = []
    for level in LEVELS:
        level_path = (
            config["hybrid_results_base"]
            / f"salt_pepper_{level}"
            / RESOLUTION
            / "results"
            / f"results_{config['dataset_name']}_hibrid_{level}.pkl"
        )
        if level_path.exists():
            all_records.extend(load_pickle(level_path.parent, level_path.name))

    save_pickle(all_records, summary_dir, all_path.name)
    save_results_to_xlsx(
        all_records,
        summary_dir,
        f"results_{config['dataset_name']}_hibrid_all.xlsx",
    )

    df = pd.DataFrame(all_records)

    aggregations = {
        "mean_ssim_nlm": ("ssim_nlm", "mean"),
        "mean_ssim_geonlm_hibrid": ("ssim_geonlm_hibrid", "mean"),
        "mean_psnr_nlm": ("psnr_nlm", "mean"),
        "mean_psnr_geonlm_hibrid": ("psnr_geonlm_hibrid", "mean"),
        "mean_score_nlm": ("score_nlm", "mean"),
        "mean_score_geonlm_hibrid": ("score_geonlm_hibrid", "mean"),
        "mean_time_geonlm_hibrid": ("time_geonlm_hibrid", "mean"),
        "mean_ssim_aswmf_nlm_h001": ("ssim_aswmf_nlm_h001", "mean"),
        "mean_psnr_aswmf_nlm_h001": ("psnr_aswmf_nlm_h001", "mean"),
        "mean_score_aswmf_nlm_h001": ("score_aswmf_nlm_h001", "mean"),
        "mean_time_aswmf_nlm_h001": ("time_aswmf_nlm_h001", "mean"),
    }
    optional_means = {
        "median": ["ssim_median", "psnr_median", "score_median", "time_median"],
        "aswmf": ["ssim_aswmf_original", "psnr_aswmf_original", "score_aswmf_original"],
        "nlmedians": ["ssim_nlmedians", "psnr_nlmedians", "score_nlmedians", "time_nlmedians"],
    }
    for method, columns in optional_means.items():
        for column in columns:
            if column in df.columns:
                metric = column.split("_", 1)[0]
                aggregations[f"mean_{metric}_{method}"] = (column, "mean")

    if "delta_score_hibrid_vs_nlm" in df.columns:
        aggregations["wins_hibrid_vs_nlm"] = (
            "delta_score_hibrid_vs_nlm",
            lambda s: int((s > 0).sum()),
        )
    if "delta_score_hibrid_vs_median" in df.columns:
        aggregations["wins_hibrid_vs_median"] = (
            "delta_score_hibrid_vs_median",
            lambda s: int((s > 0).sum()),
        )
    if "delta_score_hibrid_vs_nlmedians" in df.columns:
        aggregations["wins_hibrid_vs_nlmedians"] = (
            "delta_score_hibrid_vs_nlmedians",
            lambda s: int((s > 0).sum()),
        )
    aggregations["wins_hibrid_vs_aswmf_nlm_h001"] = (
        "delta_score_aswmf_nlm_h001_vs_geonlm_hibrid",
        lambda s: int((s < 0).sum()),
    )
    aggregations["wins_aswmf_nlm_h001_vs_hibrid"] = (
        "delta_score_aswmf_nlm_h001_vs_geonlm_hibrid",
        lambda s: int((s > 0).sum()),
    )

    summary = df.groupby("level").agg(**aggregations).reset_index()
    summary.to_excel(
        summary_dir / f"results_{config['dataset_name']}_hibrid_summary.xlsx",
        index=False,
    )
    create_augmented_method_xlsx(
        all_records,
        summary_dir / f"Results Salt and Pepper Hibrid Method - {config['dataset_label']}.xlsx",
    )
    return summary


def better(value, reference):
    if pd.isna(value) or pd.isna(reference):
        return ""
    return "melhor" if value > reference else "pior"


def create_augmented_method_xlsx(all_records, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    header = [
        "image",
        "ssim_nlm",
        "ssim_geonlm_hibrid",
        "ssim_aswmf_nlm_h001",
        "ssim_median",
        "ssim_aswmf",
        "ssim_nlmedians",
        "psnr_nlm",
        "psnr_geonlm_hibrid",
        "psnr_aswmf_nlm_h001",
        "psnr_median",
        "psnr_aswmf",
        "psnr_nlmedians",
        "score_nlm",
        "score_geonlm_hibrid",
        "score_aswmf_nlm_h001",
        "score_median",
        "score_aswmf",
        "score_nlmedians",
        "GHNLM x ASWMF_NLM_H001",
        "nlm_h",
        "h_gnlm",
        "h_aswmf_nlm_h001",
        "time_geonlm_hibrid",
        "time_aswmf_nlm_h001",
    ]
    for level in LEVELS:
        rows = [row for row in all_records if row["level"] == level]
        if not rows:
            continue
        ws = wb.create_sheet(f"Hibrid {level.title()}")
        ws.append(header)
        for record in rows:
            ws.append([
                record["file_name"],
                record.get("ssim_nlm"),
                record.get("ssim_geonlm_hibrid"),
                record.get("ssim_aswmf_nlm_h001"),
                record.get("ssim_median"),
                record.get("ssim_aswmf_original"),
                record.get("ssim_nlmedians"),
                record.get("psnr_nlm"),
                record.get("psnr_geonlm_hibrid"),
                record.get("psnr_aswmf_nlm_h001"),
                record.get("psnr_median"),
                record.get("psnr_aswmf_original"),
                record.get("psnr_nlmedians"),
                record.get("score_nlm"),
                record.get("score_geonlm_hibrid"),
                record.get("score_aswmf_nlm_h001"),
                record.get("score_median"),
                record.get("score_aswmf_original"),
                record.get("score_nlmedians"),
                better(record.get("score_geonlm_hibrid"), record.get("score_aswmf_nlm_h001")),
                record.get("nlm_h"),
                record.get("hybrid_h"),
                record.get("aswmf_nlm_h001_h"),
                record.get("time_geonlm_hibrid"),
                record.get("time_aswmf_nlm_h001"),
            ])
        style_augmented_sheet(ws, Font, PatternFill, Alignment)
    wb.save(out_path)
    return out_path


def style_augmented_sheet(ws, Font, PatternFill, Alignment):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 16
    ws.freeze_panes = "A2"


def run_experiment(config):
    all_records = []
    for level in selected_levels(config["env_prefix"]):
        all_records.extend(run_level(config, level))

    summary_dir = config["output_base"] / "summary"
    summary_dir.mkdir(parents=True, exist_ok=True)
    save_pickle(
        all_records,
        summary_dir,
        f"results_{config['dataset_name']}_aswmf_nlm_h001_all.pkl",
    )
    save_results_to_xlsx(
        all_records,
        summary_dir,
        f"results_{config['dataset_name']}_aswmf_nlm_h001_all.xlsx",
    )

    summary = make_summary(all_records, summary_dir, config["dataset_name"])
    merged_summary = merge_summary_into_hybrid(config)
    print(summary.to_string(index=False))
    if merged_summary is not None:
        print("\nMerged hybrid summary:")
        print(merged_summary.to_string(index=False))
    return summary
